import numpy as np
import pandas as pd
import openpyxl
from astropy import table

def isundef(x):
    return (isinstance(x, float) and np.isnan(x) or
            isinstance(x, np.datetime64) and np.isnat(x))


def datetime64_to_iso(x):
    return pd.Timestamp(x).isoformat()

def table_from_excel(filename, sheet=0, *, cls=table.Table,
        date_format='iso', integer_conversions=[bool, int], 
        engine=None, **kwargs):
  
    if engine is None:
        if filename.split('.')[-1] in ('xls', 'XLS'):
            engine = 'xlrd'
        elif filename.split('.')[-1] in ('xlsx', 'XLSX'):
            engine = 'openpyxl'
            # fix pandas bug that reads all empty rows 
            wb = openpyxl.load_workbook(filename)
            nrows = kwargs.get('nrows', 999999)
            header = np.atleast_1d(kwargs.get('header', []))
            skiprows = kwargs.get('skiprows', 0)
            skipfooter = kwargs.get('skipfooter', 0)
            if isinstance(sheet, int):
                sheet = wb.sheetnames[sheet]
            ws = wb[sheet]
            max_body = ws.max_row - len(header) - skiprows - skipfooter
            kwargs['nrows'] = min(max_body, nrows)
            
    ptab = pd.read_excel(filename, sheet, engine=engine, **kwargs)

    names = []
    cols = []

    for name, col in ptab.items():

        # multi-line header
        if isinstance(name, tuple):
            name = ' '.join(name)    
        names.append(name)

        col = col.to_numpy() 
        mask = [isundef(c) for c in col]
       
        # transform all dates 
        if date_format == 'iso':
            col = [datetime64_to_iso(c) if isinstance(c, np.datetime64) else c
                        for c in col]
          
        # try to transform back to int / bool 
        col = np.ma.masked_array(col, mask=mask)

        if col.dtype.kind == 'f':
            col.data[col.mask] = 0
            for dtype in integer_conversions:
                try:
                    new_col = np.ma.masked_array(col, dtype=dtype)
                    if not any((new_col - col)[~col.mask]):
                        col = new_col
                        break
                except OverflowError:
                    pass
        
        # not sure how 'nan' gets there
        if col.dtype.kind == 'U':
            col.data[col.mask] = 'N/A'

        cols.append(col)

    return cls(cols, names=names) 

if __name__ == "__main__":
    filename = 'test.xls'
    tab = table_from_excel('test.xls')
