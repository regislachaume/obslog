from openpyxl import load_workbook

def get_sheet(filename, i):

    ws = load_workbook(filename)
    sheetname = ws.sheetnames[i]
    return ws[sheetname]

filename = '/home/lachaume/git/obslog/scripts/ESO-2.2m/programs/P107/program-ESO-2.2m-P107.xls'
ws = get_sheet(filename, 1)
